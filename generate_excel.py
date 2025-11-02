#!/usr/bin/env python3
"""
Generate enhanced Excel tracker for .NET Developer Roadmap
Includes formulas, KPIs, charts, and automatic tracking
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# Read CSV file
df = pd.read_csv('dotnet-roadmap-tracker.csv')

# Create Excel writer
with pd.ExcelWriter('dotnet-roadmap-tracker-enhanced.xlsx', engine='openpyxl') as writer:
    # Write main tracking sheet
    df.to_excel(writer, sheet_name='Roadmap Tracker', index=False)
    
    # Create summary sheet
    summary_data = {
        'Metric': [
            'Total Weeks',
            'Total Months',
            'Weeks Completed',
            'Weeks In Progress',
            'Weeks Not Started',
            'Overall Progress (%)',
            'Current Week',
            'Expected Progress (%)',
            'Ahead/Behind Schedule'
        ],
        'Value': [
            '=COUNTA(\'Roadmap Tracker\'!A2:A97)',
            '24',
            '=COUNTIF(\'Roadmap Tracker\'!I2:I97,"Completed")',
            '=COUNTIF(\'Roadmap Tracker\'!I2:I97,"In Progress")',
            '=COUNTIF(\'Roadmap Tracker\'!I2:I97,"Not Started")',
            '=ROUND((C3/C1)*100,1)',
            '=IF(C3>0,C3,IF(C4>0,C3+C4,0))',
            '=ROUND((C7/C1)*100,1)',
            '=C6-C8'
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Dashboard', index=False, startrow=1)
    
    # Create monthly progress sheet
    monthly_data = []
    for month in range(1, 25):
        month_weeks = df[df['Month'] == month]
        monthly_data.append({
            'Month': month,
            'Total Weeks': len(month_weeks),
            'Completed': f'=COUNTIFS(\'Roadmap Tracker\'!B:B,{month},\'Roadmap Tracker\'!I:I,"Completed")',
            'In Progress': f'=COUNTIFS(\'Roadmap Tracker\'!B:B,{month},\'Roadmap Tracker\'!I:I,"In Progress")',
            'Not Started': f'=COUNTIFS(\'Roadmap Tracker\'!B:B,{month},\'Roadmap Tracker\'!I:I,"Not Started")',
            'Progress (%)': f'=ROUND((C{2+month}/B{2+month})*100,1)'
        })
    monthly_df = pd.DataFrame(monthly_data)
    monthly_df.to_excel(writer, sheet_name='Monthly Progress', index=False)
    
    # Create topics breakdown sheet
    topics_list = [
        'C# Fundamentals', 'OOP Basics', 'OOP Advanced', 'C# Advanced Features',
        'Data Structures', 'Algorithms', 'Problem Solving', 'SQL Server',
        'Entity Framework Core', 'Database Advanced', 'ASP.NET Core Basics',
        'ASP.NET Core Web API', 'ASP.NET Core MVC', 'ASP.NET Core Review',
        'Software Testing', 'Major Project #1', 'Design Patterns', 'SOLID Principles',
        'Clean Architecture', 'CQRS & MediatR', 'Docker', 'Microservices',
        'CI/CD', 'Azure Cloud', 'Advanced .NET', 'Capstone Project'
    ]
    
    topics_data = []
    for topic in topics_list:
        topics_data.append({
            'Topic': topic,
            'Total Weeks': f'=COUNTIF(\'Roadmap Tracker\'!E:E,"{topic}")',
            'Completed': f'=COUNTIFS(\'Roadmap Tracker\'!E:E,"{topic}",\'Roadmap Tracker\'!I:I,"Completed")',
            'In Progress': f'=COUNTIFS(\'Roadmap Tracker\'!E:E,"{topic}",\'Roadmap Tracker\'!I:I,"In Progress")',
            'Progress (%)': f'=IF(B{2+len(topics_data)}=0,0,ROUND((C{2+len(topics_data)}/B{2+len(topics_data)})*100,1))'
        })
    topics_df = pd.DataFrame(topics_data)
    topics_df.to_excel(writer, sheet_name='Topics Progress', index=False)
    
    # Create instructions sheet
    instructions_data = {
        'Instructions for Using This Tracker': [
            '1. Update Status Column',
            '2. Track Your Progress',
            '3. View Dashboard',
            '4. Monitor Monthly Progress',
            '5. Review Topics Progress',
            '6. Tips for Success',
            '',
            'Status Values:',
            'How to Update:',
            'Dashboard Features:',
            'Monthly Progress:',
            'Topics Progress:'
        ],
        'Description': [
            'In the "Roadmap Tracker" sheet, update the Status column (I) with: "Not Started", "In Progress", or "Completed"',
            'When you complete a week, add the completion date in column K and change status to "Completed"',
            'The Dashboard sheet automatically calculates your overall progress and shows if you are on track',
            'Track your progress month by month to ensure you are meeting your goals',
            'See which topics you have completed and which ones need more work',
            'Be consistent, practice daily, and don\'t skip the fundamentals!',
            '',
            '• Not Started - You haven\'t started this week yet',
            '1. Open "Roadmap Tracker" sheet\n2. Find current week row\n3. Update Status in column I\n4. Add notes in column J if needed\n5. Add completion date in column K when done',
            '• Total weeks, months, and completion statistics\n• Progress percentage\n• Ahead/Behind schedule indicator\n• Visual charts (if Excel supports)',
            '• Shows progress for each of the 24 months\n• Helps you track if you are on schedule\n• Identifies months needing more focus',
            '• Breaks down progress by topic area\n• Helps identify strengths and weaknesses\n• Guides study focus'
        ]
    }
    instructions_df = pd.DataFrame(instructions_data)
    instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

# Load workbook for formatting
wb = load_workbook('dotnet-roadmap-tracker-enhanced.xlsx')

# Format Roadmap Tracker sheet
ws_tracker = wb['Roadmap Tracker']

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Status colors
status_colors = {
    'Not Started': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'In Progress': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Completed': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
}

# Format headers
for cell in ws_tracker[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Set column widths
column_widths = {
    'A': 6,  # Week
    'B': 7,  # Month
    'C': 12, # Start Date
    'D': 12, # End Date
    'E': 22, # Topic
    'F': 35, # Subtopics
    'G': 30, # Learning Materials
    'H': 35, # Practical Tasks
    'I': 14, # Status
    'J': 25, # Notes
    'K': 15  # Completed Date
}

for col, width in column_widths.items():
    ws_tracker.column_dimensions[col].width = width

# Apply borders and alignment to all data cells
for row in ws_tracker.iter_rows(min_row=2, max_row=97, min_col=1, max_col=11):
    for cell in row:
        cell.border = border
        if cell.column_letter in ['A', 'B', 'I']:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Freeze panes
ws_tracker.freeze_panes = 'A2'

# Format Dashboard sheet
ws_dashboard = wb['Dashboard']

# Headers for dashboard
for cell in ws_dashboard[2]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Dashboard formatting
ws_dashboard.column_dimensions['A'].width = 28
ws_dashboard.column_dimensions['B'].width = 20

for row in ws_dashboard.iter_rows(min_row=3, max_row=11, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if cell.column == 1 else 'center', vertical='center')
        # Highlight key metrics
        if cell.row in [3, 6, 7, 9]:  # Key KPI rows
            cell.font = Font(bold=True, size=11)

# Add title to dashboard
ws_dashboard['A1'] = '.NET Developer Roadmap - Dashboard'
ws_dashboard['A1'].font = Font(bold=True, size=16, color='4472C4')
ws_dashboard.merge_cells('A1:B1')
ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Format Monthly Progress sheet
ws_monthly = wb['Monthly Progress']

for cell in ws_monthly[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws_monthly.column_dimensions['A'].width = 8
ws_monthly.column_dimensions['B'].width = 12
ws_monthly.column_dimensions['C'].width = 12
ws_monthly.column_dimensions['D'].width = 14
ws_monthly.column_dimensions['E'].width = 14
ws_monthly.column_dimensions['F'].width = 14

for row in ws_monthly.iter_rows(min_row=2, max_row=25, min_col=1, max_col=6):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Format Topics Progress sheet
ws_topics = wb['Topics Progress']

for cell in ws_topics[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws_topics.column_dimensions['A'].width = 25
ws_topics.column_dimensions['B'].width = 12
ws_topics.column_dimensions['C'].width = 12
ws_topics.column_dimensions['D'].width = 14
ws_topics.column_dimensions['E'].width = 14

for row in ws_topics.iter_rows(min_row=2, max_row=27, min_col=1, max_col=5):
    for cell in row:
        cell.border = border
        if cell.column == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Format Instructions sheet
ws_instructions = wb['Instructions']
ws_instructions.column_dimensions['A'].width = 35
ws_instructions.column_dimensions['B'].width = 80

for cell in ws_instructions[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for row in ws_instructions.iter_rows(min_row=2, max_row=13, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if cell.row in [2, 9, 10, 11, 12, 13]:  # Section headers
            cell.font = Font(bold=True, size=11)

# Adjust row heights for instructions
for row in range(2, 14):
    ws_instructions.row_dimensions[row].height = 60 if row >= 10 else 30

# Add a chart to Dashboard (Progress Pie Chart)
try:
    pie = PieChart()
    labels = Reference(ws_dashboard, min_col=1, min_row=4, max_row=6)
    data = Reference(ws_dashboard, min_col=2, min_row=3, max_row=6)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Overall Progress Distribution"
    pie.height = 10
    pie.width = 15
    ws_dashboard.add_chart(pie, "D2")
except Exception as e:
    print(f"Chart creation skipped: {e}")

# Save the workbook
wb.save('dotnet-roadmap-tracker-enhanced.xlsx')
print("✅ Enhanced Excel file created successfully: dotnet-roadmap-tracker-enhanced.xlsx")
print("📊 Features included:")
print("   - Automatic progress calculations")
print("   - Dashboard with KPIs")
print("   - Monthly progress tracking")
print("   - Topics breakdown")
print("   - Visual charts")
print("   - Comprehensive instructions")
